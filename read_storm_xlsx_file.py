import openpyxl

work_book = openpyxl.load_workbook('G:\\Lab\\storm_data\\storm.xlsx')

work_sheet = work_book.get_sheet_by_name('Allstorms.ibtracs_wmo.v03r09')

count_file = 1
path = 'G:\\Lab\\storm_data\\new_data_storm'
print ('start!')

for i in range(3, 196687):
	path_write = path + '\\' + str(count_file) + '.txt'
	f = open(path_write, 'a')
	if (work_sheet.cell(row = i, column = 1).value == work_sheet.cell(row = i+1, column = 1).value):
		f.write(str(work_sheet.cell(row = i, column = 5).value) + ' ' + str(work_sheet.cell(row = i, column = 8).value) + ' ' 
							+ str(work_sheet.cell(row = i, column = 9).value) + ' ' + str(work_sheet.cell(row = i, column = 10).value) + ' ' 
							+ str(work_sheet.cell(row = i, column = 13).value))
	
		f.write('\n')
		
	else:
		f.write(str(work_sheet.cell(row = i, column = 5).value) + ' ' + str(work_sheet.cell(row = i, column = 8).value) + ' ' 
							+ str(work_sheet.cell(row = i, column = 9).value) + ' ' + str(work_sheet.cell(row = i, column = 10).value) + ' ' 
							+ str(work_sheet.cell(row = i, column = 13).value))
						
		f.write('\n')				
						
		count_file = count_file + 1;
		f.close()			
	
print ('end')	




